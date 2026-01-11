"""
Journal Entry Validator - AI-Style Feedback
Reads your Excel workbook and provides detailed feedback on each exercise
"""

import openpyxl
import sys
from datetime import datetime

class JournalEntryValidator:
    def __init__(self, workbook_path):
        self.filename = workbook_path
        try:
            self.wb = openpyxl.load_workbook(workbook_path)
        except FileNotFoundError:
            print(f"❌ Error: Could not find {workbook_path}")
            print("Make sure you're in the day1_excel folder!")
            sys.exit(1)
        
        # Load Chart of Accounts
        self.coa = self.load_chart_of_accounts()
        
        self.scores = {}
        self.total_score = 0
        self.max_score = 0
    
    def load_chart_of_accounts(self):
        """Load valid account numbers"""
        ws_coa = self.wb["Chart of Accounts"]
        coa = {}
        for row in ws_coa.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Account number
                coa[str(row[0])] = {
                    'name': row[1],
                    'type': row[2],
                    'normal_balance': row[3]
                }
        return coa
    
    def validate_exercise(self, sheet_name, expected_config):
        """Validate a journal entry exercise"""
        print("\n" + "="*80)
        print(f"Validating: {sheet_name}")
        print("="*80)
        
        try:
            ws = self.wb[sheet_name]
        except KeyError:
            print(f"❌ Sheet '{sheet_name}' not found")
            return 0, 10
        
        score = 0
        max_score = 10
        feedback = []
        
        # Get entry metadata
        entry_id = ws['B3'].value
        date = ws['E3'].value
        description = ws['B4'].value
        
        print(f"\nEntry ID: {entry_id}")
        print(f"Date: {date}")
        print(f"Description: {description}\n")
        
        # Extract entry lines
        lines = []
        for row in range(7, 17):  # Lines 1-10
            account = ws[f'B{row}'].value
            debit_val = ws[f'D{row}'].value
            credit_val = ws[f'E{row}'].value
            
            # Handle formulas and convert to float
            try:
                debit = float(debit_val) if debit_val else 0
            except (ValueError, TypeError):
                debit = 0
            
            try:
                credit = float(credit_val) if credit_val else 0
            except (ValueError, TypeError):
                credit = 0
            
            if account and (debit or credit):
                lines.append({
                    'account': str(account) if account else None,
                    'debit': debit,
                    'credit': credit,
                    'row': row
                })
        
        if not lines:
            print("❌ NO ENTRY CREATED")
            print("You need to fill in account numbers and amounts!")
            return 0, max_score
        
        # Display what they created
        print(f"{'Account':<30} {'Debit':>15} {'Credit':>15}")
        print("-"*60)
        for line in lines:
            account_name = self.coa.get(line['account'], {}).get('name', 'UNKNOWN')
            debit_str = f"${line['debit']:>12,.2f}" if line['debit'] else ""
            credit_str = f"${line['credit']:>12,.2f}" if line['credit'] else ""
            print(f"{account_name:<30} {debit_str:>15} {credit_str:>15}")
        
        total_debits = sum(line['debit'] for line in lines)
        total_credits = sum(line['credit'] for line in lines)
        print("-"*60)
        print(f"{'TOTALS':<30} ${total_debits:>12,.2f} ${total_credits:>12,.2f}\n")
        
        # Validation checks
        # Check 1: Is it balanced?
        if abs(total_debits - total_credits) < 0.01:
            print("✓ BALANCED: Debits = Credits")
            score += 3
            feedback.append("Entry is properly balanced")
        else:
            print(f"❌ UNBALANCED: Debits ${total_debits:,.2f} != Credits ${total_credits:,.2f}")
            feedback.append("Entry must balance - debits must equal credits")
        
        # Check 2: Valid account numbers?
        invalid_accounts = []
        for line in lines:
            if line['account'] not in self.coa:
                invalid_accounts.append(line['account'])
        
        if not invalid_accounts:
            print("✓ VALID ACCOUNTS: All account numbers exist in Chart of Accounts")
            score += 2
            feedback.append("All accounts are valid")
        else:
            print(f"❌ INVALID ACCOUNTS: {', '.join(invalid_accounts)}")
            print("Check the Chart of Accounts sheet for valid account numbers")
            feedback.append(f"Invalid account numbers: {', '.join(invalid_accounts)}")
        
        # Check 3: Correct accounts used? (exercise-specific)
        if 'expected_accounts' in expected_config:
            expected_accounts = expected_config['expected_accounts']
            accounts_used = [line['account'] for line in lines]
            
            correct_accounts = all(acc in accounts_used for acc in expected_accounts)
            if correct_accounts:
                print(f"✓ CORRECT ACCOUNTS: You used the right accounts!")
                score += 3
                feedback.append("Correct accounts for this scenario")
            else:
                print(f"⚠️  EXPECTED ACCOUNTS: {', '.join([self.coa[acc]['name'] for acc in expected_accounts if acc in self.coa])}")
                print(f"   YOU USED: {', '.join([self.coa.get(acc, {}).get('name', acc) for acc in accounts_used])}")
                feedback.append(f"Hint: This scenario needs {', '.join([self.coa[acc]['name'] for acc in expected_accounts if acc in self.coa])}")
        
        # Check 4: Correct amounts?
        if 'expected_amount' in expected_config:
            expected = expected_config['expected_amount']
            if abs(total_debits - expected) < 0.01:
                print(f"✓ CORRECT AMOUNT: ${expected:,.2f}")
                score += 2
                feedback.append("Amount is correct")
            else:
                print(f"⚠️  EXPECTED AMOUNT: ${expected:,.2f}, YOU ENTERED: ${total_debits:,.2f}")
                feedback.append(f"Check the scenario - should total ${expected:,.2f}")
        
        # Check 5: Proper debit/credit sides?
        for line in lines:
            if line['account'] in self.coa:
                account_info = self.coa[line['account']]
                normal_balance = account_info['normal_balance']
                
                # Check if debits/credits make sense
                if normal_balance == 'Debit' and line['credit'] > 0:
                    if account_info['type'] not in ['Contra-Asset']:
                        feedback.append(f"⚠️  {account_info['name']} normally has debit balance, you credited it")
                elif normal_balance == 'Credit' and line['debit'] > 0:
                    if account_info['type'] not in ['Expense', 'Asset']:
                        feedback.append(f"⚠️  {account_info['name']} normally has credit balance, you debited it")
        
        # Provide interview tip
        print("\n" + "-"*80)
        print("💡 INTERVIEW TIP:")
        if score >= 8:
            print("   EXCELLENT! You could explain this entry confidently in an interview.")
            print(f"   'I created an entry to {description.lower()}...'")
        elif score >= 5:
            print("   GOOD START! Review the feedback and try again.")
        else:
            print("   NEEDS WORK. Review accrual accounting basics.")
        
        print("\n" + "-"*80)
        print(f"SCORE: {score}/{max_score}")
        
        if feedback:
            print("\nFeedback:")
            for fb in feedback:
                print(f"  • {fb}")
        
        return score, max_score
    
    def validate_all(self):
        """Validate all exercises"""
        print("\n" + "="*80)
        print("JOURNAL ENTRY WORKBOOK VALIDATION")
        print("="*80)
        print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        exercises = [
            {
                'name': 'Exercise 1 - Maintenance',
                'expected_accounts': ['6100', '2100'],
                'expected_amount': 12500
            },
            {
                'name': 'Exercise 2 - Prepaid',
                'expected_accounts': ['1200', '6300'],
                'expected_amount': 5000
            },
            {
                'name': 'Exercise 3 - Debug',
                'expected_accounts': ['6700', '1600'],
                'expected_amount': 3500  # Should be 3500 debit, 3500 credit after fix
            },
            {
                'name': 'Exercise 4 - Payroll',
                'expected_accounts': ['6600', '2200', '2300', '1000'],
                'expected_amount': 45000
            },
            {
                'name': 'Exercise 5 - Month End',
                'expected_accounts': ['6200', '6700', '6300', '6800', '2100', '1600', '1200'],
                'expected_amount': 14450  # 8200 + 4500 + 1000 + 750
            },
        ]
        
        for ex in exercises:
            score, max_score = self.validate_exercise(ex['name'], ex)
            self.scores[ex['name']] = (score, max_score)
            self.total_score += score
            self.max_score += max_score
        
        # Final summary
        self.print_summary()
        
        # Write results back to Excel
        self.write_scores_to_excel()
    
    def print_summary(self):
        """Print final summary"""
        print("\n\n" + "="*80)
        print("FINAL SUMMARY")
        print("="*80)
        
        for exercise, (score, max_score) in self.scores.items():
            pct = int(score/max_score*100) if max_score > 0 else 0
            status = "✓" if pct >= 70 else "⚠️" if pct >= 50 else "❌"
            print(f"{status} {exercise:<35} {score:>2}/{max_score} ({pct}%)")
        
        print("-"*80)
        total_pct = int(self.total_score/self.max_score*100) if self.max_score > 0 else 0
        print(f"{'TOTAL SCORE':<35} {self.total_score:>2}/{self.max_score} ({total_pct}%)\n")
        
        if total_pct >= 80:
            print("🏆 EXCELLENT WORK! You're ready for reconciliations.")
            print("   Next: py validate_reconciliation.py bank_reconciliation.xlsx")
        elif total_pct >= 60:
            print("💪 GOOD PROGRESS! Review any failed exercises and try again.")
        else:
            print("📚 KEEP PRACTICING! Review accounting fundamentals.")
            print("   Remember: Expenses are debits, Liabilities/Revenue are credits")
        
        print("="*80)
    
    def write_scores_to_excel(self):
        """Write scores back to the Validation sheet in Excel"""
        try:
            from openpyxl.styles import Font, PatternFill
            
            # Try to find Validation sheet
            if "Validation" not in self.wb.sheetnames:
                return  # No validation sheet, skip
            
            ws = self.wb["Validation"]
            
            # Write scores starting at row 6 (where exercises are listed)
            row_mapping = {
                'Exercise 1 - Maintenance': 6,
                'Exercise 2 - Prepaid': 7,
                'Exercise 3 - Debug': 8,
                'Exercise 4 - Payroll': 9,
                'Exercise 5 - Month End': 10,
            }
            
            for exercise, (score, max_score) in self.scores.items():
                if exercise in row_mapping:
                    row = row_mapping[exercise]
                    pct = int(score/max_score*100) if max_score > 0 else 0
                    
                    # Column B: Score
                    ws[f'B{row}'] = f"{score}/{max_score}"
                    ws[f'B{row}'].font = Font(bold=True, size=11)
                    
                    # Column C: Percentage
                    ws[f'C{row}'] = f"{pct}%"
                    ws[f'C{row}'].font = Font(bold=True, size=11)
                    
                    # Column D: Status with color
                    if pct >= 70:
                        ws[f'D{row}'] = "✓ PASS"
                        ws[f'D{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        ws[f'D{row}'].font = Font(color="006100", bold=True)
                    elif pct >= 50:
                        ws[f'D{row}'] = "⚠️ REVIEW"
                        ws[f'D{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        ws[f'D{row}'].font = Font(color="9C5700", bold=True)
                    else:
                        ws[f'D{row}'] = "❌ REDO"
                        ws[f'D{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        ws[f'D{row}'].font = Font(color="9C0006", bold=True)
            
            # Write total score
            total_pct = int(self.total_score/self.max_score*100) if self.max_score > 0 else 0
            ws['A12'] = "TOTAL:"
            ws['A12'].font = Font(bold=True, size=12)
            ws['B12'] = f"{self.total_score}/{self.max_score}"
            ws['B12'].font = Font(bold=True, size=12)
            ws['C12'] = f"{total_pct}%"
            ws['C12'].font = Font(bold=True, size=12)
            
            if total_pct >= 80:
                ws['D12'] = "🏆 EXCELLENT!"
                ws['D12'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws['D12'].font = Font(color="006100", bold=True, size=12)
            elif total_pct >= 60:
                ws['D12'] = "💪 GOOD"
                ws['D12'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                ws['D12'].font = Font(color="9C5700", bold=True, size=12)
            else:
                ws['D12'] = "📚 PRACTICE MORE"
                ws['D12'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws['D12'].font = Font(color="9C0006", bold=True, size=12)
            
            # Save the workbook
            self.wb.save(self.filename)
            print(f"\n✓ Scores written to Validation sheet in {self.filename}")
            
        except Exception as e:
            print(f"\n⚠️  Could not write scores to Excel: {e}")

def main():
    if len(sys.argv) < 2:
        print("Usage: py validate_journal_entries.py journal_entry_workbook.xlsx")
        sys.exit(1)
    
    workbook_path = sys.argv[1]
    validator = JournalEntryValidator(workbook_path)
    validator.validate_all()

if __name__ == "__main__":
    main()



