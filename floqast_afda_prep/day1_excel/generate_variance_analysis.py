"""
Generate Variance Analysis Template - Realistic P&L Analysis
This creates an Excel file with pivot tables and variance analysis
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import random

def create_variance_analysis():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # === SHEET 1: MONTHLY P&L DATA ===
    ws_data = wb.create_sheet("Monthly P&L Data")
    
    ws_data['A1'] = "Thind Transport - Monthly P&L Data"
    ws_data['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws_data['A2'] = "December 2024 vs. November 2024"
    ws_data['A2'].font = Font(size=11, italic=True)
    
    # Headers
    headers = ["Account", "Category", "November Actual", "December Actual", "Budget (Dec)", "Variance vs Nov", "Variance vs Budget", "% Change"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # P&L data
    pl_data = [
        # Account, Category, Nov, Dec, Budget
        ["Revenue", "Revenue", 285000, 312000, 300000],
        ["Cost of Revenue", "COGS", -125000, -145000, -135000],
        ["Gross Profit", "Gross Profit", 160000, 167000, 165000],
        ["Payroll Expense", "Operating Expense", -45000, -47000, -45000],
        ["Fuel Expense", "Operating Expense", -22000, -28500, -25000],
        ["Maintenance Expense", "Operating Expense", -15000, -21000, -18000],
        ["Insurance Expense", "Operating Expense", -6000, -6000, -6000],
        ["Office Expense", "Operating Expense", -3500, -4200, -4000],
        ["Professional Fees", "Operating Expense", -2500, -3000, -2500],
        ["Depreciation Expense", "Operating Expense", -4500, -4500, -4500],
        ["Interest Expense", "Operating Expense", -750, -750, -750],
        ["Total Operating Expenses", "Operating Expense", -99250, -114950, -105750],
        ["Net Income", "Net Income", 60750, 52050, 59250],
    ]
    
    for row_idx, row_data in enumerate(pl_data, start=5):
        # Account name
        ws_data.cell(row=row_idx, column=1, value=row_data[0])
        ws_data.cell(row=row_idx, column=2, value=row_data[1])
        
        # November Actual
        cell_nov = ws_data.cell(row=row_idx, column=3, value=row_data[2])
        cell_nov.number_format = '#,##0'
        if row_data[2] < 0:
            cell_nov.font = Font(color="C00000")
        
        # December Actual
        cell_dec = ws_data.cell(row=row_idx, column=4, value=row_data[3])
        cell_dec.number_format = '#,##0'
        if row_data[3] < 0:
            cell_dec.font = Font(color="C00000")
        
        # Budget
        cell_bud = ws_data.cell(row=row_idx, column=5, value=row_data[4])
        cell_bud.number_format = '#,##0'
        if row_data[4] < 0:
            cell_bud.font = Font(color="C00000")
        
        # Formulas for variances
        # Variance vs Nov (Dec - Nov)
        var_nov_cell = ws_data.cell(row=row_idx, column=6, value=f"=D{row_idx}-C{row_idx}")
        var_nov_cell.number_format = '#,##0'
        
        # Variance vs Budget (Dec - Budget)
        var_bud_cell = ws_data.cell(row=row_idx, column=7, value=f"=D{row_idx}-E{row_idx}")
        var_bud_cell.number_format = '#,##0'
        
        # % Change vs Nov
        pct_cell = ws_data.cell(row=row_idx, column=8, value=f"=IF(C{row_idx}=0,0,(D{row_idx}-C{row_idx})/ABS(C{row_idx}))")
        pct_cell.number_format = '0.0%'
        
        # Bold totals
        if "Total" in row_data[0] or "Net Income" in row_data[0] or "Gross Profit" in row_data[0]:
            for col in range(1, 9):
                ws_data.cell(row=row_idx, column=col).font = Font(bold=True)
                ws_data.cell(row=row_idx, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Column widths
    ws_data.column_dimensions['A'].width = 30
    ws_data.column_dimensions['B'].width = 20
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws_data.column_dimensions[col].width = 15
    ws_data.column_dimensions['H'].width = 12
    
    # === SHEET 2: VARIANCE ANALYSIS ===
    ws_analysis = wb.create_sheet("Variance Analysis")
    
    ws_analysis['A1'] = "VARIANCE ANALYSIS WORKSHEET"
    ws_analysis['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws_analysis['A2'] = "Thind Transport - December 2024"
    ws_analysis['A2'].font = Font(size=11, italic=True)
    
    # Instructions
    ws_analysis['A4'] = "YOUR TASK:"
    ws_analysis['A4'].font = Font(bold=True, size=11, color="C00000")
    ws_analysis['A5'] = "1. Review the Monthly P&L Data sheet"
    ws_analysis['A6'] = "2. Identify material variances (>10% change or >$5,000 variance)"
    ws_analysis['A7'] = "3. For each material variance, provide explanation in the table below"
    ws_analysis['A8'] = "4. This mimics what FloQast's Flux AI agent would do automatically"
    
    # Variance table header
    headers = ["Account", "December Actual", "Benchmark", "Variance $", "Variance %", "Material?", "Your Explanation"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_analysis.cell(row=10, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Key variances to analyze
    variances = [
        ["Revenue", "='Monthly P&L Data'!D5", "='Monthly P&L Data'!C5", "='Monthly P&L Data'!F5", "='Monthly P&L Data'!H5"],
        ["Fuel Expense", "='Monthly P&L Data'!D9", "='Monthly P&L Data'!C9", "='Monthly P&L Data'!F9", "='Monthly P&L Data'!H9"],
        ["Maintenance Expense", "='Monthly P&L Data'!D10", "='Monthly P&L Data'!C10", "='Monthly P&L Data'!F10", "='Monthly P&L Data'!H10"],
        ["Net Income", "='Monthly P&L Data'!D17", "='Monthly P&L Data'!C17", "='Monthly P&L Data'!F17", "='Monthly P&L Data'!H17"],
    ]
    
    for row_idx, var_data in enumerate(variances, start=11):
        ws_analysis.cell(row=row_idx, column=1, value=var_data[0])
        
        for col_idx in range(2, 6):
            cell = ws_analysis.cell(row=row_idx, column=col_idx, value=var_data[col_idx-1])
            if col_idx == 5:  # Percentage column
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0'
        
        # Material? formula (Yes if >10% or >$5000)
        material_cell = ws_analysis.cell(row=row_idx, column=6, value=f"=IF(OR(ABS(D{row_idx})>5000,ABS(E{row_idx})>0.1),\"YES\",\"NO\")")
        
        # Explanation placeholder
        ws_analysis.cell(row=row_idx, column=7, value="[Your explanation here]")
    
    # Column widths
    ws_analysis.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D']:
        ws_analysis.column_dimensions[col].width = 15
    ws_analysis.column_dimensions['E'].width = 12
    ws_analysis.column_dimensions['F'].width = 12
    ws_analysis.column_dimensions['G'].width = 50
    
    # Example explanations
    ws_analysis['A16'] = "EXAMPLE EXPLANATIONS:"
    ws_analysis['A16'].font = Font(bold=True, size=11, color="1F4E78")
    ws_analysis['A17'] = "• Revenue increased 9.5% due to 3 new customer contracts starting in December"
    ws_analysis['A18'] = "• Fuel Expense increased 29.5% due to seasonal peak demand + fuel price spike (avg $4.50/gal vs $3.80)"
    ws_analysis['A19'] = "• Maintenance increased 40% due to scheduled annual fleet inspection ($10K) + unexpected repairs ($6K)"
    
    # === SHEET 3: FLUX SIMULATION ===
    ws_flux = wb.create_sheet("Flux AI Simulation")
    
    ws_flux['A1'] = "FloQast Flux AI - Automated Variance Explanations"
    ws_flux['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws_flux['A2'] = "This is what FloQast's AI would generate automatically"
    ws_flux['A2'].font = Font(size=11, italic=True)
    
    ws_flux['A4'] = "Material Variances Detected:"
    ws_flux['A4'].font = Font(bold=True, size=12)
    
    # Simulated AI explanations
    ai_explanations = [
        ["Revenue", "$27,000 increase (9.5%)", "Revenue exceeded prior month by $27K due to increased shipment volume. December saw 15% more deliveries (285 vs 247 in November). New customer contracts: ABC Corp ($12K), XYZ Ltd ($8K), Johnson Co ($7K)."],
        ["Fuel Expense", "$6,500 increase (29.5%)", "Fuel costs spiked 29.5% above November. Contributing factors: 1) 15% increase in mileage (peak season), 2) Fuel price increase from $3.80/gal to $4.50/gal (18.4% price inflation). Total: 2,850 gallons vs 2,400 in November."],
        ["Maintenance Expense", "$6,000 increase (40%)", "Maintenance 40% above November due to scheduled annual fleet inspection ($10K) performed Dec 15-20. Additionally, unexpected transmission repairs on 3 trucks ($6K total). Remainder is routine maintenance."],
        ["Net Income", "$8,700 decrease (14.3%)", "Net income declined 14.3% despite revenue growth, driven by: 1) Higher COGS due to increased volume, 2) Fuel cost spike, 3) Scheduled maintenance timing. Normalized for one-time fleet inspection, core profitability remains strong."],
    ]
    
    row = 6
    for account, variance, explanation in ai_explanations:
        ws_flux[f'A{row}'] = f"🔍 {account}"
        ws_flux[f'A{row}'].font = Font(bold=True, size=11, color="1F4E78")
        ws_flux[f'B{row}'] = variance
        ws_flux[f'B{row}'].font = Font(bold=True)
        
        ws_flux[f'A{row+1}'] = explanation
        ws_flux.merge_cells(f'A{row+1}:G{row+1}')
        ws_flux[f'A{row+1}'].alignment = Alignment(wrap_text=True)
        
        row += 3
    
    ws_flux.column_dimensions['A'].width = 80
    
    # === SHEET 4: VALIDATION ===
    ws_val = wb.create_sheet("Validation Results")
    
    ws_val['A1'] = "VALIDATION DASHBOARD"
    ws_val['A1'].font = Font(size=14, bold=True)
    ws_val['A3'] = "Run validation script: py validate_variance_analysis.py variance_analysis.xlsx"
    ws_val['A5'] = "Results will appear here after validation."
    
    # Save workbook
    wb.save('variance_analysis.xlsx')
    print("✓ Created variance_analysis.xlsx")

if __name__ == "__main__":
    create_variance_analysis()

