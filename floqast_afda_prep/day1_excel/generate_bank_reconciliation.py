"""
Generate Bank Reconciliation Template - Realistic Excel Worksheet
This creates an Excel file that mimics real bank reconciliation workflows
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta
import random

def create_bank_reconciliation():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # === SHEET 1: GL CASH TRANSACTIONS ===
    ws_gl = wb.create_sheet("GL Cash Transactions")
    
    ws_gl['A1'] = "GENERAL LEDGER - CASH ACCOUNT (1000)"
    ws_gl['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws_gl['A2'] = "Thind Transport - December 2024"
    ws_gl['A2'].font = Font(size=11, italic=True)
    
    # Headers
    headers = ["Trans ID", "Date", "Description", "Ref", "Amount", "Type"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_gl.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # GL transactions data
    gl_data = [
        ["GL-1001", "12/02/2024", "Customer Payment - ABC Corp", "INV-5001", 15000, "Receipt"],
        ["GL-1002", "12/05/2024", "Fuel Purchase - Express Gas", "CHK-2001", -850, "Payment"],
        ["GL-1003", "12/08/2024", "Customer Payment - XYZ Ltd", "INV-5002", 22000, "Receipt"],
        ["GL-1004", "12/12/2024", "Office Supplies - Depot", "CHK-2002", -450, "Payment"],
        ["GL-1005", "12/15/2024", "Customer Payment - Smith Co", "INV-5003", 18500, "Receipt"],
        ["GL-1006", "12/20/2024", "Maintenance - Fleet Services", "CHK-2003", -3200, "Payment"],
        ["GL-1007", "12/28/2024", "Check #2004 - Insurance", "CHK-2004", -5000, "Payment"],
        ["GL-1008", "12/29/2024", "Customer Payment - Johnson", "INV-5004", 8500, "Receipt"],
        ["GL-1009", "12/30/2024", "Customer Payment - Brown Inc", "INV-5005", 12000, "Receipt"],
    ]
    
    for row_idx, row_data in enumerate(gl_data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_gl.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left' if col_idx <= 4 else 'right', vertical='center')
            if col_idx == 5:  # Amount column
                cell.number_format = '#,##0.00'
                if value < 0:
                    cell.font = Font(color="C00000")  # Red for negative
    
    # Column widths
    ws_gl.column_dimensions['A'].width = 12
    ws_gl.column_dimensions['B'].width = 12
    ws_gl.column_dimensions['C'].width = 35
    ws_gl.column_dimensions['D'].width = 12
    ws_gl.column_dimensions['E'].width = 15
    ws_gl.column_dimensions['F'].width = 12
    
    # === SHEET 2: BANK STATEMENT ===
    ws_bank = wb.create_sheet("Bank Statement")
    
    ws_bank['A1'] = "BANK STATEMENT - First National Bank"
    ws_bank['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws_bank['A2'] = "Account: ****1234 - Thind Transport"
    ws_bank['A2'].font = Font(size=11, italic=True)
    ws_bank['A3'] = "Statement Period: December 1-31, 2024"
    ws_bank['A3'].font = Font(size=11, italic=True)
    
    # Headers
    headers = ["Trans ID", "Date", "Description", "Amount", "Type"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_bank.cell(row=5, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Bank statement data (some match GL, some don't)
    bank_data = [
        ["BNK-A001", "12/02/2024", "Deposit - ABC Corp", 15000, "Deposit"],
        ["BNK-A002", "12/06/2024", "Debit - Express Gas", -850, "Payment"],  # Date differs by 1 day
        ["BNK-A003", "12/08/2024", "Deposit - XYZ Ltd", 22000, "Deposit"],
        ["BNK-A004", "12/12/2024", "Debit - Office Depot", -450, "Payment"],
        ["BNK-A005", "12/15/2024", "Deposit - Smith Co", 18500, "Deposit"],
        ["BNK-A006", "12/22/2024", "Debit - Fleet Services", -3200, "Payment"],  # Date differs
        ["BNK-A007", "12/15/2024", "Bank Service Charge", -25, "Fee"],  # Not in GL
        # Note: CHK-2004 (Insurance) not on bank yet - Outstanding check
        # Note: Recent deposits (12/29, 12/30) not on bank yet - Deposits in transit
    ]
    
    for row_idx, row_data in enumerate(bank_data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_bank.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left' if col_idx <= 3 else 'right', vertical='center')
            if col_idx == 4:  # Amount column
                cell.number_format = '#,##0.00'
                if value < 0:
                    cell.font = Font(color="C00000")
    
    # Column widths
    ws_bank.column_dimensions['A'].width = 12
    ws_bank.column_dimensions['B'].width = 12
    ws_bank.column_dimensions['C'].width = 35
    ws_bank.column_dimensions['D'].width = 15
    ws_bank.column_dimensions['E'].width = 12
    
    # === SHEET 3: RECONCILIATION WORKSHEET ===
    ws_recon = wb.create_sheet("Reconciliation Worksheet")
    
    ws_recon['A1'] = "BANK RECONCILIATION WORKSHEET"
    ws_recon['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws_recon['A2'] = "Thind Transport - December 31, 2024"
    ws_recon['A2'].font = Font(size=11, italic=True)
    
    # Instructions
    ws_recon['A4'] = "YOUR TASK:"
    ws_recon['A4'].font = Font(bold=True, size=11, color="C00000")
    ws_recon['A5'] = "1. Review GL Cash Transactions and Bank Statement sheets"
    ws_recon['A6'] = "2. In column G below, mark 'MATCHED' for transactions that appear on both"
    ws_recon['A7'] = "3. Identify deposits in transit (in GL but not on bank statement)"
    ws_recon['A8'] = "4. Identify outstanding checks (in GL but not cleared bank)"
    ws_recon['A9'] = "5. Identify bank charges (on bank but not in GL)"
    ws_recon['A10'] = "6. Complete reconciliation summary at bottom"
    
    # GL Transactions section
    ws_recon['A12'] = "GL TRANSACTIONS:"
    ws_recon['A12'].font = Font(bold=True, size=11)
    
    headers = ["GL ID", "Date", "Description", "Amount", "Bank ID", "Bank Date", "Match Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_recon.cell(row=13, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Copy GL data
    for row_idx, row_data in enumerate(gl_data, start=14):
        ws_recon.cell(row=row_idx, column=1, value=row_data[0])  # GL ID
        ws_recon.cell(row=row_idx, column=2, value=row_data[1])  # Date
        ws_recon.cell(row=row_idx, column=3, value=row_data[2])  # Description
        cell = ws_recon.cell(row=row_idx, column=4, value=row_data[4])  # Amount
        cell.number_format = '#,##0.00'
        if row_data[4] < 0:
            cell.font = Font(color="C00000")
        
        # Columns E, F, G are for you to fill in
        ws_recon.cell(row=row_idx, column=5, value="[Fill bank ID]")
        ws_recon.cell(row=row_idx, column=6, value="[Fill bank date]")
        ws_recon.cell(row=row_idx, column=7, value="[MATCHED/DIT/OC]")
    
    # Column widths
    ws_recon.column_dimensions['A'].width = 12
    ws_recon.column_dimensions['B'].width = 12
    ws_recon.column_dimensions['C'].width = 35
    ws_recon.column_dimensions['D'].width = 15
    ws_recon.column_dimensions['E'].width = 12
    ws_recon.column_dimensions['F'].width = 12
    ws_recon.column_dimensions['G'].width = 18
    
    # Reconciliation summary - TWO COLUMN FORMAT
    summary_row = 25
    ws_recon[f'A{summary_row}'] = "RECONCILIATION SUMMARY:"
    ws_recon[f'A{summary_row}'].font = Font(bold=True, size=12, color="1F4E78")
    
    # LEFT SIDE: Bank Statement Balance
    ws_recon[f'A{summary_row+2}'] = "BANK STATEMENT SIDE:"
    ws_recon[f'A{summary_row+2}'].font = Font(bold=True, size=11, color="1F4E78")
    
    ws_recon[f'A{summary_row+3}'] = "Bank Statement Balance (12/31):"
    ws_recon[f'C{summary_row+3}'] = "[Calculate from Bank Statement sheet]"
    
    ws_recon[f'A{summary_row+4}'] = "Add: Deposits in Transit"
    ws_recon[f'C{summary_row+4}'] = "[Sum of DIT transactions]"
    
    ws_recon[f'A{summary_row+5}'] = "Less: Outstanding Checks"
    ws_recon[f'C{summary_row+5}'] = "[Sum of OC transactions]"
    
    ws_recon[f'A{summary_row+7}'] = "Adjusted Bank Balance:"
    ws_recon[f'C{summary_row+7}'] = "[Calculate above]"
    ws_recon[f'C{summary_row+7}'].font = Font(bold=True, size=11)
    ws_recon[f'C{summary_row+7}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # RIGHT SIDE: GL Balance
    ws_recon[f'E{summary_row+2}'] = "GL BOOK SIDE:"
    ws_recon[f'E{summary_row+2}'].font = Font(bold=True, size=11, color="1F4E78")
    
    ws_recon[f'E{summary_row+3}'] = "GL Cash Balance (12/31):"
    ws_recon[f'G{summary_row+3}'] = "=SUM('GL Cash Transactions'!E5:E13)"
    ws_recon[f'G{summary_row+3}'].number_format = '#,##0.00'
    
    ws_recon[f'E{summary_row+4}'] = "Less: Bank Charges Not in GL"
    ws_recon[f'G{summary_row+4}'] = "[Enter amount: -25]"
    
    ws_recon[f'E{summary_row+5}'] = "Add: Interest Income (if any)"
    ws_recon[f'G{summary_row+5}'] = "0"
    ws_recon[f'G{summary_row+5}'].number_format = '#,##0.00'
    
    ws_recon[f'E{summary_row+7}'] = "Adjusted GL Balance:"
    ws_recon[f'G{summary_row+7}'] = "[Calculate above]"
    ws_recon[f'G{summary_row+7}'].font = Font(bold=True, size=11)
    ws_recon[f'G{summary_row+7}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Match check
    ws_recon[f'A{summary_row+9}'] = "RESULT:"
    ws_recon[f'A{summary_row+9}'].font = Font(bold=True, size=12, color="C00000")
    ws_recon[f'A{summary_row+10}'] = "If both adjusted balances equal $66,475, you're CORRECT!"
    ws_recon[f'A{summary_row+10}'].font = Font(italic=True, color="70AD47")
    
    # Column widths for summary
    ws_recon.column_dimensions['E'].width = 28
    ws_recon.column_dimensions['G'].width = 20
    
    # === SHEET 4: VALIDATION ===
    ws_val = wb.create_sheet("Validation Results")
    
    ws_val['A1'] = "VALIDATION DASHBOARD"
    ws_val['A1'].font = Font(size=14, bold=True)
    ws_val['A3'] = "Run validation script: py validate_reconciliation.py bank_reconciliation.xlsx"
    ws_val['A5'] = "Results will appear here after validation."
    
    # Save workbook
    wb.save('bank_reconciliation.xlsx')
    print("Created bank_reconciliation.xlsx successfully")

if __name__ == "__main__":
    create_bank_reconciliation()



