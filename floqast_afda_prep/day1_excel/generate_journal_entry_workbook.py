"""
Generate Journal Entry Workbook - Realistic Excel Template
This creates an Excel file that mimics real ERP journal entry screens
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_journal_entry_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # === SHEET 1: CHART OF ACCOUNTS ===
    ws_coa = wb.create_sheet("Chart of Accounts")
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Headers
    ws_coa['A1'] = "Account Number"
    ws_coa['B1'] = "Account Name"
    ws_coa['C1'] = "Type"
    ws_coa['D1'] = "Normal Balance"
    
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws_coa[cell].fill = header_fill
        ws_coa[cell].font = header_font
        ws_coa[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Chart of Accounts data
    coa_data = [
        ["1000", "Cash", "Asset", "Debit"],
        ["1100", "Accounts Receivable", "Asset", "Debit"],
        ["1200", "Prepaid Insurance", "Asset", "Debit"],
        ["1500", "Equipment", "Asset", "Debit"],
        ["1600", "Accumulated Depreciation", "Contra-Asset", "Credit"],
        ["2000", "Accounts Payable", "Liability", "Credit"],
        ["2100", "Accrued Liabilities", "Liability", "Credit"],
        ["2200", "Payroll Tax Payable", "Liability", "Credit"],
        ["2300", "Employee Benefits Payable", "Liability", "Credit"],
        ["3000", "Retained Earnings", "Equity", "Credit"],
        ["4000", "Revenue", "Revenue", "Credit"],
        ["6100", "Maintenance Expense", "Expense", "Debit"],
        ["6200", "Fuel Expense", "Expense", "Debit"],
        ["6300", "Insurance Expense", "Expense", "Debit"],
        ["6400", "Professional Fees", "Expense", "Debit"],
        ["6500", "Office Expense", "Expense", "Debit"],
        ["6600", "Payroll Expense", "Expense", "Debit"],
        ["6700", "Depreciation Expense", "Expense", "Debit"],
        ["6800", "Interest Expense", "Expense", "Debit"],
    ]
    
    for idx, row_data in enumerate(coa_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_coa.cell(row=idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto-fit columns
    ws_coa.column_dimensions['A'].width = 15
    ws_coa.column_dimensions['B'].width = 30
    ws_coa.column_dimensions['C'].width = 15
    ws_coa.column_dimensions['D'].width = 15
    
    # === SHEET 2: JE TEMPLATE (Master Template) ===
    ws_template = wb.create_sheet("JE Template")
    
    # Title
    ws_template.merge_cells('A1:F1')
    ws_template['A1'] = "JOURNAL ENTRY FORM"
    ws_template['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws_template['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_template['A1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws_template.row_dimensions[1].height = 30
    
    # Entry metadata
    ws_template['A3'] = "Entry ID:"
    ws_template['B3'] = "JE-XXXXX-XXX"
    ws_template['D3'] = "Date:"
    ws_template['E3'] = "MM/DD/YYYY"
    
    ws_template['A4'] = "Description:"
    ws_template.merge_cells('B4:F4')
    ws_template['B4'] = "[Enter description of journal entry]"
    
    # Make metadata bold
    for cell in ['A3', 'D3', 'A4']:
        ws_template[cell].font = Font(bold=True)
    
    # Column headers for entry lines
    ws_template['A6'] = "Line"
    ws_template['B6'] = "Account Number"
    ws_template['C6'] = "Account Name"
    ws_template['D6'] = "Debit"
    ws_template['E6'] = "Credit"
    ws_template['F6'] = "Notes"
    
    for cell in ['A6', 'B6', 'C6', 'D6', 'E6', 'F6']:
        ws_template[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_template[cell].font = Font(color="FFFFFF", bold=True)
        ws_template[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Entry lines (10 lines for entries)
    for line_num in range(1, 11):
        row = 6 + line_num
        ws_template[f'A{row}'] = line_num
        ws_template[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Format number cells
        ws_template[f'D{row}'].number_format = '#,##0.00'
        ws_template[f'E{row}'].number_format = '#,##0.00'
    
    # Totals row
    totals_row = 17
    ws_template[f'A{totals_row}'] = "TOTALS:"
    ws_template[f'A{totals_row}'].font = Font(bold=True, size=12)
    ws_template[f'D{totals_row}'] = f'=SUM(D7:D16)'
    ws_template[f'E{totals_row}'] = f'=SUM(E7:E16)'
    ws_template[f'D{totals_row}'].number_format = '#,##0.00'
    ws_template[f'E{totals_row}'].number_format = '#,##0.00'
    ws_template[f'D{totals_row}'].font = Font(bold=True)
    ws_template[f'E{totals_row}'].font = Font(bold=True)
    ws_template[f'D{totals_row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws_template[f'E{totals_row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Balance check
    ws_template['A19'] = "Status:"
    ws_template['A19'].font = Font(bold=True, size=12)
    ws_template['B19'] = f'=IF(ABS(D{totals_row}-E{totals_row})<0.01,"✓ BALANCED","✗ UNBALANCED")'
    ws_template['B19'].font = Font(size=12, bold=True)
    
    # Conditional formatting for balance check (manually set color based on balance)
    ws_template['B19'].alignment = Alignment(horizontal='left')
    
    # Column widths
    ws_template.column_dimensions['A'].width = 8
    ws_template.column_dimensions['B'].width = 18
    ws_template.column_dimensions['C'].width = 30
    ws_template.column_dimensions['D'].width = 15
    ws_template.column_dimensions['E'].width = 15
    ws_template.column_dimensions['F'].width = 25
    
    # === EXERCISE 1: MAINTENANCE ACCRUAL ===
    ws_ex1 = wb.create_sheet("Exercise 1 - Maintenance")
    copy_template_to_exercise(ws_ex1, ws_template)
    
    ws_ex1['B3'] = "JE-MAINT-001"
    ws_ex1['E3'] = "12/31/2024"
    ws_ex1['B4'] = "Fleet maintenance accrual - December"
    
    # Instructions
    ws_ex1['A21'] = "INSTRUCTIONS:"
    ws_ex1['A21'].font = Font(bold=True, size=11, color="C00000")
    ws_ex1['A22'] = "Scenario: Fleet maintenance done in December for $12,500, invoice arrives in January."
    ws_ex1['A23'] = "Your Task: Create the accrual journal entry. Fill in Account Numbers, Debits, and Credits."
    ws_ex1['A24'] = "Hint: Expenses increase with debits, liabilities increase with credits."
    
    # === EXERCISE 2: PREPAID INSURANCE ===
    ws_ex2 = wb.create_sheet("Exercise 2 - Prepaid")
    copy_template_to_exercise(ws_ex2, ws_template)
    
    ws_ex2['B3'] = "JE-PREP-001"
    ws_ex2['E3'] = "12/31/2024"
    ws_ex2['B4'] = "Prepaid insurance adjustment"
    
    ws_ex2['A21'] = "INSTRUCTIONS:"
    ws_ex2['A21'].font = Font(bold=True, size=11, color="C00000")
    ws_ex2['A22'] = "Scenario: Paid $6,000 for 6 months insurance on Dec 1. Only 1 month used. Record prepaid."
    ws_ex2['A23'] = "Your Task: Calculate prepaid amount ($5,000) and create adjustment entry."
    ws_ex2['A24'] = "Hint: Prepaid Insurance is an asset (debit), Insurance Expense is expense (debit)"
    
    # === EXERCISE 3: DEBUG DEPRECIATION ===
    ws_ex3 = wb.create_sheet("Exercise 3 - Debug")
    copy_template_to_exercise(ws_ex3, ws_template)
    
    ws_ex3['B3'] = "JE-DEPR-001"
    ws_ex3['E3'] = "12/31/2024"
    ws_ex3['B4'] = "Monthly depreciation"
    
    # Pre-fill with WRONG entry
    ws_ex3['B7'] = "6700"
    ws_ex3['C7'] = "Depreciation Expense"
    ws_ex3['D7'] = 3500
    ws_ex3['B8'] = "1600"
    ws_ex3['C8'] = "Accumulated Depreciation"
    ws_ex3['D8'] = 3000  # WRONG: should be credit
    
    ws_ex3['A21'] = "INSTRUCTIONS:"
    ws_ex3['A21'].font = Font(bold=True, size=11, color="C00000")
    ws_ex3['A22'] = "Challenge: This entry is BROKEN. Find and fix the error(s)."
    ws_ex3['A23'] = "Hint: Check if debits = credits. Check if accounts have correct normal balances."
    
    # === EXERCISE 4: PAYROLL ===
    ws_ex4 = wb.create_sheet("Exercise 4 - Payroll")
    copy_template_to_exercise(ws_ex4, ws_template)
    
    ws_ex4['B3'] = "JE-PAYROLL-001"
    ws_ex4['E3'] = "12/31/2024"
    ws_ex4['B4'] = "December payroll"
    
    ws_ex4['A21'] = "INSTRUCTIONS:"
    ws_ex4['A21'].font = Font(bold=True, size=11, color="C00000")
    ws_ex4['A22'] = "Scenario: Gross wages $45,000. Withholdings: $6,800 taxes, $2,200 benefits. Net pay: $36,000."
    ws_ex4['A23'] = "Your Task: Create 4-line compound entry (1 debit for expense, 3 credits for payables/cash)."
    ws_ex4['A24'] = "Accounts: 6600-Payroll Expense, 2200-Payroll Tax Payable, 2300-Benefits Payable, 1000-Cash"
    
    # === EXERCISE 5: FULL MONTH-END ===
    ws_ex5 = wb.create_sheet("Exercise 5 - Month End")
    
    # This one is blank - they need to create multiple entries
    ws_ex5['A1'] = "EXERCISE 5: FULL MONTH-END CLOSE"
    ws_ex5['A1'].font = Font(size=14, bold=True, color="1F4E78")
    
    ws_ex5['A3'] = "INSTRUCTIONS:"
    ws_ex5['A3'].font = Font(bold=True, size=11, color="C00000")
    ws_ex5['A4'] = "Create FOUR separate journal entries for Thind Transport's December close:"
    ws_ex5['A5'] = "1. Fuel accrual: $8,200 in fuel charges, invoice pending"
    ws_ex5['A6'] = "2. Depreciation: $4,500 monthly depreciation"
    ws_ex5['A7'] = "3. Insurance expense: $1,000 monthly insurance (from prepaid)"
    ws_ex5['A8'] = "4. Interest accrual: $750 interest on loan"
    ws_ex5['A10'] = "Use the JE Template sheet as a guide. Create each entry below."
    
    # === VALIDATION SHEET ===
    ws_val = wb.create_sheet("Validation")
    
    ws_val['A1'] = "VALIDATION DASHBOARD"
    ws_val['A1'].font = Font(size=14, bold=True)
    ws_val['A3'] = "Run validation script: py validate_journal_entries.py journal_entry_workbook.xlsx"
    ws_val['A5'] = "Exercises:"
    ws_val['A6'] = "Exercise 1 - Maintenance Accrual"
    ws_val['A7'] = "Exercise 2 - Prepaid Insurance"
    ws_val['A8'] = "Exercise 3 - Debug Depreciation"
    ws_val['A9'] = "Exercise 4 - Payroll"
    ws_val['A10'] = "Exercise 5 - Month-End Close"
    
    ws_val['B5'] = "Status"
    ws_val['C5'] = "Score"
    ws_val['B5'].font = Font(bold=True)
    ws_val['C5'].font = Font(bold=True)
    
    # Save workbook
    wb.save('journal_entry_workbook.xlsx')
    print("✓ Created journal_entry_workbook.xlsx")

def copy_template_to_exercise(target_ws, template_ws):
    """Copy template layout to exercise sheet"""
    # Copy values and formatting from template
    for row in template_ws.iter_rows(min_row=1, max_row=19, min_col=1, max_col=6):
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column)
            
            # Copy value
            if cell.value and not isinstance(cell.value, str) or (isinstance(cell.value, str) and not cell.value.startswith('=')):
                target_cell.value = cell.value
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                # Copy formula
                target_cell.value = cell.value
            
            # Copy formatting
            if cell.has_style:
                target_cell.font = cell.font.copy()
                target_cell.fill = cell.fill.copy()
                target_cell.alignment = cell.alignment.copy()
                target_cell.number_format = cell.number_format
    
    # Copy column dimensions
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        target_ws.column_dimensions[col].width = template_ws.column_dimensions[col].width
    
    # Copy row dimensions
    target_ws.row_dimensions[1].height = template_ws.row_dimensions[1].height

if __name__ == "__main__":
    create_journal_entry_workbook()



