"""
Alaska Airlines Treasury Financial Analyst - Excel Practice Simulation Builder
Creates a realistic 6-tab workbook simulating Treasury operations
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

def create_treasury_simulation():
    """Create the complete Treasury practice simulation workbook"""
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create all tabs
    create_instructions_tab(wb)
    create_fleet_lease_portfolio_tab(wb)
    create_debt_amortization_tab(wb)
    create_cash_flow_forecast_tab(wb)
    create_payment_reconciliation_tab(wb)
    create_variance_report_tab(wb)
    create_month_end_close_tab(wb)
    
    # Save workbook
    filename = 'Alaska_Treasury_Practice_Simulation.xlsx'
    wb.save(filename)
    print(f"[SUCCESS] Created: {filename}")
    return filename

def set_header_style(cell):
    """Apply consistent header styling"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def set_title_style(cell):
    """Apply title styling"""
    cell.font = Font(bold=True, size=14, color="1F4E78")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def create_instructions_tab(wb):
    """Tab 0: Instructions and Overview"""
    ws = wb.create_sheet("Instructions", 0)
    
    # Title
    ws['A1'] = "ALASKA AIRLINES TREASURY FINANCIAL ANALYST"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws['A2'] = "Excel Practice Simulation - Work Simulation"
    ws['A2'].font = Font(bold=True, size=14, color="1F4E78")
    
    ws['A4'] = "OVERVIEW"
    set_title_style(ws['A4'])
    
    instructions = [
        ("", ""),
        ("Purpose:", "This workbook simulates real Treasury Financial Analyst tasks at an airline."),
        ("", "You'll work with aircraft financing, debt portfolios, lease payments, reconciliations,"),
        ("", "and month-end close packages - exactly what Alaska's Treasury team does daily."),
        ("", ""),
        ("Time:", "Complete all exercises in 2-3 hours. Don't worry about perfection on first try."),
        ("", ""),
        ("Structure:", "6 Exercises across 6 tabs, building on each other"),
        ("", ""),
        ("EXERCISES", ""),
        ("", ""),
        ("Tab 1:", "FLEET & LEASE PORTFOLIO"),
        ("", "• Review Alaska's aircraft portfolio (5 planes with different financing)"),
        ("", "• Build a summary dashboard showing total monthly obligations by aircraft"),
        ("", "• Calculate total portfolio monthly payment"),
        ("Difficulty:", "⭐ Beginner - Warmup exercise"),
        ("", ""),
        ("Tab 2:", "DEBT AMORTIZATION SCHEDULE"),
        ("", "• Build complete amortization schedule for $50M aircraft financing"),
        ("", "• Calculate monthly payment, principal/interest split, remaining balance"),
        ("", "• Use Excel functions: PMT, IPMT, PPMT"),
        ("Difficulty:", "⭐⭐ Intermediate - Core Treasury skill"),
        ("", ""),
        ("Tab 3:", "MONTHLY CASH FLOW FORECAST"),
        ("", "• Build 12-month rolling forecast of ALL debt and lease obligations"),
        ("", "• Pull from Tabs 1 and 2, consolidate total monthly outflows"),
        ("", "• Flag months with high payment concentration using conditional formatting"),
        ("Difficulty:", "⭐⭐⭐ Advanced - Requires linking across tabs"),
        ("", ""),
        ("Tab 4:", "PAYMENT TRACKING & RECONCILIATION"),
        ("", "• Reconcile bank statement (8 wire transfers) against payment schedule"),
        ("", "• Use VLOOKUP or INDEX/MATCH to match payments"),
        ("", "• Identify and categorize variances (timing, amount, missing payments)"),
        ("Difficulty:", "⭐⭐ Intermediate - Daily Treasury task"),
        ("", ""),
        ("Tab 5:", "VARIANCE REPORT"),
        ("", "• Analyze forecasted vs. actual payments by category"),
        ("", "• Build pivot table to summarize variances by type"),
        ("", "• Create chart showing forecast accuracy"),
        ("", "• Flag variances > 5% for investigation"),
        ("Difficulty:", "⭐⭐ Intermediate - Monthly reporting"),
        ("", ""),
        ("Tab 6:", "MONTH-END CLOSE PACKAGE"),
        ("", "• Prepare accounting close package with all key metrics"),
        ("", "• Calculate beginning/ending debt balance, interest expense, lease expense"),
        ("", "• Format as journal entry (Debit/Credit format)"),
        ("", "• Build summary dashboard for leadership"),
        ("Difficulty:", "⭐⭐⭐ Advanced - Comprehensive exercise"),
        ("", ""),
        ("TIPS FOR SUCCESS", ""),
        ("", ""),
        ("1.", "Start with Tab 1 and work sequentially - later tabs build on earlier ones"),
        ("2.", "Read the scenario and task carefully on each tab before starting"),
        ("3.", "Use named ranges or cell references when linking between tabs"),
        ("4.", "Add notes/comments to document your approach"),
        ("5.", "If stuck, move on and come back - you can complete exercises independently"),
        ("6.", "Check the Practice Guide (separate file) for hints and Treasury terminology"),
        ("", ""),
        ("WHAT YOU'LL LEARN", ""),
        ("", ""),
        ("✓", "Debt amortization calculations (PMT, IPMT, PPMT functions)"),
        ("✓", "Payment tracking and reconciliation (VLOOKUP, INDEX/MATCH)"),
        ("✓", "Cash flow forecasting across multiple debt instruments"),
        ("✓", "Variance analysis and reporting"),
        ("✓", "Month-end close package preparation"),
        ("✓", "Treasury terminology and workflows"),
        ("✓", "How to talk about Treasury work in interviews"),
        ("", ""),
        ("INTERVIEW VALUE", ""),
        ("", ""),
        ("After completing this simulation, you can say in interviews:", ""),
        ("", '• "I built a practice Treasury workbook to prepare for this role"'),
        ("", '• "I\'m familiar with debt amortization schedules and payment tracking"'),
        ("", '• "I understand how Treasury supports month-end close with accounting"'),
        ("", '• "I practiced reconciliation workflows similar to what Treasury does daily"'),
        ("", ""),
        ("This shows initiative, genuine interest, and that you understand the role", ""),
        ("beyond just reading the job description.", ""),
        ("", ""),
        ("Ready? Start with Tab 1: Fleet & Lease Portfolio →", ""),
    ]
    
    row = 5
    for col_a, col_b in instructions:
        ws[f'A{row}'] = col_a
        ws[f'B{row}'] = col_b
        
        if col_a in ["EXERCISES", "TIPS FOR SUCCESS", "WHAT YOU'LL LEARN", "INTERVIEW VALUE"]:
            ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
        elif "Tab " in col_a:
            ws[f'A{row}'].font = Font(bold=True, color="1F4E78")
        elif col_a in ["✓", "1.", "2.", "3.", "4.", "5.", "6."]:
            ws[f'A{row}'].font = Font(bold=True, color="E67E22")
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 90

def create_fleet_lease_portfolio_tab(wb):
    """Tab 1: Fleet & Lease Portfolio"""
    ws = wb.create_sheet("1_Fleet_Portfolio")
    
    # Title
    ws['A1'] = "TAB 1: FLEET & LEASE PORTFOLIO"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    
    ws['A3'] = "SCENARIO"
    set_title_style(ws['A3'])
    ws['A4'] = "You're the Financial Analyst supporting Alaska Airlines Treasury. The team manages financing"
    ws['A5'] = "for the company's aircraft fleet. You need to understand the current portfolio and monthly obligations."
    
    ws['A7'] = "YOUR TASK"
    set_title_style(ws['A7'])
    ws['A8'] = "1. Review the aircraft portfolio data below"
    ws['A9'] = "2. Calculate the TOTAL monthly obligation for each aircraft (see note about financed aircraft)"
    ws['A10'] = "3. Build a summary showing total portfolio monthly payment"
    ws['A11'] = "4. Identify which aircraft has the highest monthly cost"
    
    ws['A13'] = "ALASKA AIRLINES AIRCRAFT PORTFOLIO"
    set_title_style(ws['A13'])
    
    # Headers
    headers = ['Aircraft ID', 'Aircraft Type', 'Financing Type', 'Delivery Date', 
               'Monthly Lease Payment', 'Monthly Debt Payment', 'Total Monthly Obligation']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=14, column=col_num)
        cell.value = header
        set_header_style(cell)
    
    # Aircraft data
    aircraft_data = [
        ['AA-001', 'Boeing 737 MAX 9', 'Operating Lease', '2023-01-15', 385000, '-', ''],
        ['AA-002', 'Boeing 737-900ER', 'Finance Lease', '2021-06-01', '-', 425000, ''],
        ['AA-003', 'Airbus A321neo', 'Operating Lease', '2024-03-20', 410000, '-', ''],
        ['AA-004', 'Boeing 737 MAX 8', 'Finance Lease', '2022-11-10', '-', 398000, ''],
        ['AA-005', 'Boeing 737-800', 'Operating Lease', '2020-08-05', 295000, '-', ''],
    ]
    
    for row_num, row_data in enumerate(aircraft_data, 15):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if col_num in [5, 6]:  # Payment columns
                if value == '-':
                    cell.value = '-'
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.value = value
                    cell.number_format = '$#,##0'
            else:
                cell.value = value
    
    # Instructions box
    ws['A21'] = "NOTE:"
    ws['A21'].font = Font(bold=True, color="E74C3C")
    ws['A22'] = "• Operating Lease: Airline pays monthly lease payment (like renting)"
    ws['A23'] = "• Finance Lease: Airline is buying the aircraft and pays monthly debt payment"
    ws['A24'] = "• For Total Monthly Obligation: Use the payment that applies (lease OR debt, not both)"
    ws['A25'] = "• Hint: Use an IF statement or simple formula to fill column G"
    
    ws['A27'] = "SUMMARY DASHBOARD (Build This)"
    set_title_style(ws['A27'])
    ws['A28'] = "Total Monthly Portfolio Payment:"
    ws['A28'].font = Font(bold=True)
    ws['B28'] = "[Calculate total of column G]"
    ws['B28'].font = Font(italic=True, color="7F7F7F")
    
    ws['A29'] = "Highest Cost Aircraft:"
    ws['A29'].font = Font(bold=True)
    ws['B29'] = "[Use MAX function to find]"
    ws['B29'].font = Font(italic=True, color="7F7F7F")
    
    ws['A31'] = "CHALLENGE (Optional):"
    ws['A31'].font = Font(bold=True, color="E67E22")
    ws['A32'] = "• Calculate the percentage each aircraft represents of total portfolio cost"
    ws['A33'] = "• Use conditional formatting to highlight aircraft with payments > $400,000/month"
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 25

def create_debt_amortization_tab(wb):
    """Tab 2: Debt Amortization Schedule"""
    ws = wb.create_sheet("2_Debt_Amortization")
    
    # Title
    ws['A1'] = "TAB 2: DEBT AMORTIZATION SCHEDULE"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    
    ws['A3'] = "SCENARIO"
    set_title_style(ws['A3'])
    ws['A4'] = "Alaska Airlines financed aircraft AA-002 (Boeing 737-900ER) with a $50,000,000 loan."
    ws['A5'] = "You need to build a complete amortization schedule to track principal and interest payments."
    
    ws['A7'] = "YOUR TASK"
    set_title_style(ws['A7'])
    ws['A8'] = "1. Calculate the monthly payment amount using the PMT function"
    ws['A9'] = "2. Build a 144-row amortization schedule (12 years x 12 months)"
    ws['A10'] = "3. Calculate principal and interest components for each payment"
    ws['A11'] = "4. Track the remaining loan balance after each payment"
    ws['A12'] = "5. Verify that the final balance equals $0"
    
    ws['A14'] = "LOAN PARAMETERS"
    set_title_style(ws['A14'])
    
    # Loan details box
    ws['A15'] = "Loan Amount:"
    ws['A15'].font = Font(bold=True)
    ws['B15'] = 50000000
    ws['B15'].number_format = '$#,##0'
    
    ws['A16'] = "Annual Interest Rate:"
    ws['A16'].font = Font(bold=True)
    ws['B16'] = 0.055
    ws['B16'].number_format = '0.00%'
    
    ws['A17'] = "Loan Term (Years):"
    ws['A17'].font = Font(bold=True)
    ws['B17'] = 12
    
    ws['A18'] = "Loan Term (Months):"
    ws['A18'].font = Font(bold=True)
    ws['B18'] = 144
    
    ws['A19'] = "Monthly Interest Rate:"
    ws['A19'].font = Font(bold=True)
    ws['B19'] = "=B16/12"
    ws['B19'].number_format = '0.0000%'
    
    ws['A21'] = "Monthly Payment:"
    ws['A21'].font = Font(bold=True, color="E74C3C")
    ws['B21'] = "[Use PMT function here]"
    ws['B21'].font = Font(italic=True, color="7F7F7F")
    ws['C21'] = "Hint: =PMT(rate, nper, pv)"
    ws['C21'].font = Font(italic=True, size=9, color="7F7F7F")
    
    # Amortization schedule headers
    ws['A24'] = "AMORTIZATION SCHEDULE (Build Below)"
    set_title_style(ws['A24'])
    
    headers = ['Payment #', 'Payment Date', 'Beginning Balance', 'Monthly Payment', 
               'Interest Paid', 'Principal Paid', 'Ending Balance']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=25, column=col_num)
        cell.value = header
        set_header_style(cell)
    
    # First 3 rows as examples
    ws['A26'] = 1
    ws['B26'] = datetime(2021, 7, 1)
    ws['B26'].number_format = 'yyyy-mm-dd'
    ws['C26'] = "=$B$15"
    ws['D26'] = "=$B$21"
    ws['E26'] = "[Use IPMT function]"
    ws['E26'].font = Font(italic=True, color="7F7F7F")
    ws['F26'] = "[Use PPMT function]"
    ws['F26'].font = Font(italic=True, color="7F7F7F")
    ws['G26'] = "[Beginning Balance - Principal Paid]"
    ws['G26'].font = Font(italic=True, color="7F7F7F")
    
    ws['A27'] = 2
    ws['B27'] = datetime(2021, 8, 1)
    ws['B27'].number_format = 'yyyy-mm-dd'
    ws['C27'] = "[Previous Ending Balance]"
    ws['C27'].font = Font(italic=True, color="7F7F7F")
    
    ws['A28'] = 3
    ws['B28'] = datetime(2021, 9, 1)
    ws['B28'].number_format = 'yyyy-mm-dd'
    
    ws['A29'] = "..."
    ws['A29'].font = Font(italic=True, color="7F7F7F")
    
    ws['A31'] = "HINTS:"
    ws['A31'].font = Font(bold=True, color="E67E22")
    ws['A32'] = "• IPMT(rate, per, nper, pv) - calculates interest portion of payment"
    ws['A33'] = "• PPMT(rate, per, nper, pv) - calculates principal portion of payment"
    ws['A34'] = "• rate = monthly interest rate (B19), per = payment number, nper = total payments (B18), pv = loan amount (B15)"
    ws['A35'] = "• Ending Balance = Beginning Balance - Principal Paid"
    ws['A36'] = "• Payment Date: Add 1 month each row (use =DATE or =EDATE functions)"
    
    ws['A38'] = "VERIFICATION:"
    ws['A38'].font = Font(bold=True, color="1F4E78")
    ws['A39'] = "Final Payment (Row 169):"
    ws['A40'] = "Ending Balance should be $0 (or very close due to rounding)"
    ws['A41'] = "Total Interest Paid:"
    ws['A41'].font = Font(bold=True)
    ws['B41'] = "[Sum all Interest Paid]"
    ws['B41'].font = Font(italic=True, color="7F7F7F")
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18

def create_cash_flow_forecast_tab(wb):
    """Tab 3: Monthly Cash Flow Forecast"""
    ws = wb.create_sheet("3_Cash_Flow_Forecast")
    
    # Title
    ws['A1'] = "TAB 3: MONTHLY CASH FLOW FORECAST"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    
    ws['A3'] = "SCENARIO"
    set_title_style(ws['A3'])
    ws['A4'] = "Treasury needs to forecast the company's monthly cash outflows for debt and lease obligations"
    ws['A5'] = "over the next 12 months. This helps ensure sufficient liquidity and plan for upcoming payments."
    
    ws['A7'] = "YOUR TASK"
    set_title_style(ws['A7'])
    ws['A8'] = "1. Pull total monthly lease payments from Tab 1 (Fleet Portfolio)"
    ws['A9'] = "2. Pull monthly debt payment from Tab 2 (for the financed aircraft)"
    ws['A10'] = "3. Create a 12-month rolling forecast showing total monthly obligations"
    ws['A11'] = "4. Use conditional formatting to flag months where total payments > $2,000,000"
    ws['A12'] = "5. Calculate cumulative cash outflow over the 12-month period"
    
    ws['A14'] = "12-MONTH CASH FLOW FORECAST (Build Below)"
    set_title_style(ws['A14'])
    
    # Headers
    headers = ['Month', 'Month Date', 'Operating Lease Payments', 'Debt Service Payments', 
               'Total Monthly Outflow', 'Cumulative Outflow', 'Alert Flag']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=15, column=col_num)
        cell.value = header
        set_header_style(cell)
    
    # First few rows with hints
    start_date = datetime(2026, 2, 1)
    for i in range(12):
        row = 16 + i
        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=2).value = start_date + timedelta(days=30*i)
        ws.cell(row=row, column=2).number_format = 'mmm-yyyy'
    
    ws['C16'] = "[Link to Tab 1 summary]"
    ws['C16'].font = Font(italic=True, color="7F7F7F")
    ws['D16'] = "[Link to Tab 2 monthly payment]"
    ws['D16'].font = Font(italic=True, color="7F7F7F")
    ws['E16'] = "[Sum columns C + D]"
    ws['E16'].font = Font(italic=True, color="7F7F7F")
    ws['F16'] = "[Running total]"
    ws['F16'].font = Font(italic=True, color="7F7F7F")
    ws['G16'] = '[IF Total > $2M, "HIGH", ""]'
    ws['G16'].font = Font(italic=True, color="7F7F7F")
    
    ws['A29'] = "SUMMARY METRICS"
    set_title_style(ws['A29'])
    
    ws['A30'] = "Total 12-Month Cash Outflow:"
    ws['A30'].font = Font(bold=True)
    ws['B30'] = "[Sum all monthly outflows]"
    ws['B30'].font = Font(italic=True, color="7F7F7F")
    
    ws['A31'] = "Average Monthly Payment:"
    ws['A31'].font = Font(bold=True)
    ws['B31'] = "[Average of monthly outflows]"
    ws['B31'].font = Font(italic=True, color="7F7F7F")
    
    ws['A32'] = "Peak Month Payment:"
    ws['A32'].font = Font(bold=True)
    ws['B32'] = "[Max monthly outflow]"
    ws['B32'].font = Font(italic=True, color="7F7F7F")
    
    ws['A34'] = "HINTS:"
    ws['A34'].font = Font(bold=True, color="E67E22")
    ws['A35'] = "• Use cell references between tabs: ='1_Fleet_Portfolio'!B28"
    ws['A36'] = "• Conditional formatting: Select column E, use Format > Conditional Formatting > Highlight Cells Rules"
    ws['A37'] = "• Alert Flag: Use IF function to flag high payment months for Treasury planning"
    
    ws['A39'] = "WHY THIS MATTERS:"
    ws['A39'].font = Font(bold=True, color="1F4E78")
    ws['A40'] = "Treasury uses rolling cash flow forecasts to:"
    ws['A41'] = "  • Ensure sufficient cash on hand for upcoming obligations"
    ws['A42'] = "  • Plan for high-payment months (may need to draw on credit facilities)"
    ws['A43'] = "  • Provide visibility to Finance and Accounting teams"
    ws['A44'] = "  • Support credit rating agency reporting"
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 23
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12

def create_payment_reconciliation_tab(wb):
    """Tab 4: Payment Tracking & Reconciliation"""
    ws = wb.create_sheet("4_Payment_Reconciliation")
    
    # Title
    ws['A1'] = "TAB 4: PAYMENT TRACKING & RECONCILIATION"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    
    ws['A3'] = "SCENARIO"
    set_title_style(ws['A3'])
    ws['A4'] = "You receive the January bank statement showing 8 wire transfers for aircraft payments."
    ws['A5'] = "You need to reconcile these actual payments against the scheduled payments to verify accuracy."
    ws['A6'] = "This is a daily Treasury task - catching errors before they become problems."
    
    ws['A8'] = "YOUR TASK"
    set_title_style(ws['A8'])
    ws['A9'] = "1. Review the scheduled payments (what we expected to pay)"
    ws['A10'] = "2. Review the actual bank wire transfers (what we actually paid)"
    ws['A11'] = "3. Match each bank payment to scheduled payment using VLOOKUP or INDEX/MATCH"
    ws['A12'] = "4. Calculate variances (Actual - Scheduled)"
    ws['A13'] = "5. Categorize each variance (Timing, Amount, Missing, Extra)"
    ws['A14'] = "6. Calculate the reconciled balance"
    
    # Scheduled Payments
    ws['A16'] = "SCHEDULED PAYMENTS (January 2026)"
    set_title_style(ws['A16'])
    
    headers1 = ['Payment ID', 'Payee', 'Aircraft', 'Scheduled Amount', 'Scheduled Date', 'Payment Type']
    for col_num, header in enumerate(headers1, 1):
        cell = ws.cell(row=17, column=col_num)
        cell.value = header
        set_header_style(cell)
    
    scheduled_payments = [
        ['PAY-001', 'Boeing Capital', 'AA-002', 425000, '2026-01-05', 'Debt Service'],
        ['PAY-002', 'GECAS Leasing', 'AA-001', 385000, '2026-01-10', 'Operating Lease'],
        ['PAY-003', 'AerCap', 'AA-003', 410000, '2026-01-15', 'Operating Lease'],
        ['PAY-004', 'Wells Fargo Leasing', 'AA-004', 398000, '2026-01-20', 'Debt Service'],
        ['PAY-005', 'BBAM Leasing', 'AA-005', 295000, '2026-01-25', 'Operating Lease'],
    ]
    
    for row_num, row_data in enumerate(scheduled_payments, 18):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if col_num == 4:  # Amount
                cell.value = value
                cell.number_format = '$#,##0'
            elif col_num == 5:  # Date
                cell.value = value
                cell.number_format = 'yyyy-mm-dd'
            else:
                cell.value = value
    
    # Bank Statement
    ws['H16'] = "BANK STATEMENT (Actual Wire Transfers)"
    set_title_style(ws['H16'])
    
    headers2 = ['Wire Ref', 'Wire Date', 'Payee', 'Amount Paid', 'Reference Note']
    for col_num, header in enumerate(headers2, 8):
        cell = ws.cell(row=17, column=col_num)
        cell.value = header
        set_header_style(cell)
    
    bank_wires = [
        ['WIRE-3891', '2026-01-05', 'Boeing Capital', 425000, 'PAY-001 - AA-002'],
        ['WIRE-3892', '2026-01-10', 'GECAS Leasing', 385000, 'PAY-002 - AA-001'],
        ['WIRE-3893', '2026-01-14', 'AerCap', 410000, 'PAY-003 - AA-003'],  # Early by 1 day
        ['WIRE-3894', '2026-01-20', 'Wells Fargo Leasing', 400000, 'PAY-004 - AA-004'],  # $2K overpayment
        ['WIRE-3895', '2026-01-25', 'BBAM Leasing', 295000, 'PAY-005 - AA-005'],
        ['WIRE-3896', '2026-01-28', 'Airbus Financial', 125000, 'Deposit on new aircraft'],  # Extra payment
        # Missing: One scheduled payment not in bank statement (will be caught in recon)
    ]
    
    for row_num, row_data in enumerate(bank_wires, 18):
        for col_num, value in enumerate(row_data, 8):
            cell = ws.cell(row=row_num, column=col_num)
            if col_num == 9:  # Date
                cell.value = value
                cell.number_format = 'yyyy-mm-dd'
            elif col_num == 10:  # Amount
                cell.value = value
                cell.number_format = '$#,##0'
            else:
                cell.value = value
    
    # Reconciliation Worksheet
    ws['A25'] = "RECONCILIATION WORKSHEET (Build This)"
    set_title_style(ws['A25'])
    
    headers3 = ['Payment ID', 'Scheduled Amount', 'Actual Amount', 'Variance', 
                'Variance %', 'Status', 'Variance Category', 'Notes']
    for col_num, header in enumerate(headers3, 1):
        cell = ws.cell(row=26, column=col_num)
        cell.value = header
        set_header_style(cell)
    
    ws['A27'] = "PAY-001"
    ws['B27'] = "[Lookup from scheduled]"
    ws['B27'].font = Font(italic=True, color="7F7F7F")
    ws['C27'] = "[VLOOKUP from bank statement]"
    ws['C27'].font = Font(italic=True, color="7F7F7F")
    ws['D27'] = "[Actual - Scheduled]"
    ws['D27'].font = Font(italic=True, color="7F7F7F")
    ws['E27'] = "[Variance / Scheduled]"
    ws['E27'].font = Font(italic=True, color="7F7F7F")
    ws['F27'] = '[IF variance=0, "Match", "Variance"]'
    ws['F27'].font = Font(italic=True, color="7F7F7F")
    
    ws['A33'] = "SUMMARY"
    set_title_style(ws['A33'])
    ws['A34'] = "Total Scheduled Payments:"
    ws['A34'].font = Font(bold=True)
    ws['B34'] = "[Sum scheduled amounts]"
    ws['B34'].font = Font(italic=True, color="7F7F7F")
    
    ws['A35'] = "Total Actual Payments:"
    ws['A35'].font = Font(bold=True)
    ws['B35'] = "[Sum actual amounts]"
    ws['B35'].font = Font(italic=True, color="7F7F7F")
    
    ws['A36'] = "Net Variance:"
    ws['A36'].font = Font(bold=True)
    ws['B36'] = "[Actual - Scheduled]"
    ws['B36'].font = Font(italic=True, color="7F7F7F")
    
    ws['A38'] = "HINTS:"
    ws['A38'].font = Font(bold=True, color="E67E22")
    ws['A39'] = "• Use VLOOKUP or INDEX/MATCH to find bank payments matching each Payment ID"
    ws['A40'] = "• Look for clues in the 'Reference Note' column of bank statement"
    ws['A41'] = "• Variance Categories: 'Match', 'Timing Difference', 'Amount Difference', 'Missing Payment', 'Extra Payment'"
    ws['A42'] = "• Note: PAY-003 paid early, PAY-004 overpaid by $2K, one payment missing from bank statement"
    
    ws['A44'] = "WHY THIS MATTERS:"
    ws['A44'].font = Font(bold=True, color="1F4E78")
    ws['A45'] = "Reconciliation catches:"
    ws['A46'] = "  • Overpayments/underpayments that need correction"
    ws['A47'] = "  • Missed payments that could cause default"
    ws['A48'] = "  • Unauthorized wire transfers (fraud prevention)"
    ws['A49'] = "  • Accounting errors before month-end close"
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['L'].width = 30

def create_variance_report_tab(wb):
    """Tab 5: Variance Report"""
    ws = wb.create_sheet("5_Variance_Report")
    
    # Title
    ws['A1'] = "TAB 5: VARIANCE REPORT & ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    
    ws['A3'] = "SCENARIO"
    set_title_style(ws['A3'])
    ws['A4'] = "At month-end, Treasury prepares a variance report comparing forecasted vs. actual payments."
    ws['A5'] = "This goes to leadership and accounting to explain any differences from the forecast."
    
    ws['A7'] = "YOUR TASK"
    set_title_style(ws['A7'])
    ws['A8'] = "1. Review the forecast vs. actual data below"
    ws['A9'] = "2. Calculate variances (Actual - Forecast) and variance %"
    ws['A10'] = "3. Create a Pivot Table to summarize variances by category"
    ws['A11'] = "4. Build a chart showing forecast vs. actual by payment type"
    ws['A12'] = "5. Flag any variances > 5% for investigation"
    
    ws['A14'] = "JANUARY 2026: FORECAST VS. ACTUAL"
    set_title_style(ws['A14'])
    
    # Headers
    headers = ['Payment Category', 'Forecasted Amount', 'Actual Amount', 'Variance ($)', 
               'Variance (%)', 'Investigation Required?', 'Variance Reason']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=15, column=col_num)
        cell.value = header
        set_header_style(cell)
    
    # Variance data with some realistic discrepancies
    variance_data = [
        ['Debt Service - AA-002', 425000, 425000, '', '', '', 'On target'],
        ['Debt Service - AA-004', 398000, 400000, '', '', '', 'Overpayment - investigating with lender'],
        ['Operating Lease - AA-001', 385000, 385000, '', '', '', 'On target'],
        ['Operating Lease - AA-003', 410000, 410000, '', '', '', 'On target (paid 1 day early)'],
        ['Operating Lease - AA-005', 295000, 295000, '', '', '', 'On target'],
        ['Interest on Credit Facility', 45000, 47800, '', '', '', 'Higher interest rates this month'],
        ['Lease Security Deposits', 0, 125000, '', '', '', 'New aircraft deposit - not in forecast'],
        ['FX Impact on Leases', 0, -8500, '', '', '', 'EUR/USD favorable movement'],
    ]
    
    for row_num, row_data in enumerate(variance_data, 16):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if col_num in [2, 3]:  # Amount columns
                cell.value = value
                cell.number_format = '$#,##0'
            else:
                cell.value = value
    
    ws['D16'] = "[Actual - Forecast]"
    ws['D16'].font = Font(italic=True, color="7F7F7F")
    ws['E16'] = "[Variance / Forecast]"
    ws['E16'].font = Font(italic=True, color="7F7F7F")
    ws['F16'] = '[IF ABS(variance%) > 5%, "YES", "NO"]'
    ws['F16'].font = Font(italic=True, color="7F7F7F")
    
    ws['A25'] = "SUMMARY ANALYSIS"
    set_title_style(ws['A25'])
    
    ws['A26'] = "Total Forecasted:"
    ws['A26'].font = Font(bold=True)
    ws['B26'] = "[Sum forecast column]"
    ws['B26'].font = Font(italic=True, color="7F7F7F")
    
    ws['A27'] = "Total Actual:"
    ws['A27'].font = Font(bold=True)
    ws['B27'] = "[Sum actual column]"
    ws['B27'].font = Font(italic=True, color="7F7F7F")
    
    ws['A28'] = "Total Variance:"
    ws['A28'].font = Font(bold=True)
    ws['B28'] = "[Sum variance column]"
    ws['B28'].font = Font(italic=True, color="7F7F7F")
    
    ws['A29'] = "Forecast Accuracy:"
    ws['A29'].font = Font(bold=True)
    ws['B29'] = "[1 - ABS(Total Variance / Total Forecast)]"
    ws['B29'].font = Font(italic=True, color="7F7F7F")
    
    ws['A31'] = "Items Requiring Investigation:"
    ws['A31'].font = Font(bold=True)
    ws['B31'] = "[Count of 'YES' in column F]"
    ws['B31'].font = Font(italic=True, color="7F7F7F")
    
    ws['A33'] = "PIVOT TABLE EXERCISE"
    set_title_style(ws['A33'])
    ws['A34'] = "Create a Pivot Table to summarize:"
    ws['A35'] = "  • Total variance by payment type (Debt Service vs. Operating Lease vs. Other)"
    ws['A36'] = "  • Count of variances by investigation status"
    ws['A37'] = "  • Place the pivot table starting in cell H15"
    
    ws['A39'] = "CHART EXERCISE"
    set_title_style(ws['A39'])
    ws['A40'] = "Create a column chart comparing Forecast vs. Actual:"
    ws['A41'] = "  • X-axis: Payment categories"
    ws['A42'] = "  • Y-axis: Dollar amounts"
    ws['A43'] = "  • Two series: Forecasted (blue) and Actual (orange)"
    
    ws['A45'] = "HINTS:"
    ws['A45'].font = Font(bold=True, color="E67E22")
    ws['A46'] = "• Variance % formula: =(Actual - Forecast) / Forecast"
    ws['A47'] = "• Use ABS() function for absolute value when checking if > 5%"
    ws['A48'] = "• Conditional formatting: Highlight variances requiring investigation in red"
    ws['A49'] = "• Pivot Table: Insert > PivotTable, select data range A15:G23"
    
    ws['A51'] = "WHY THIS MATTERS:"
    ws['A51'].font = Font(bold=True, color="1F4E78")
    ws['A52'] = "Variance reporting shows:"
    ws['A53'] = "  • Forecast accuracy (how well Treasury predicted cash needs)"
    ws['A54'] = "  • Unexpected items that impact liquidity"
    ws['A55'] = "  • Areas needing process improvement"
    ws['A56'] = "  • Explanations for month-end accounting variance"
    
    # Set column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 40

def create_month_end_close_tab(wb):
    """Tab 6: Month-End Close Package"""
    ws = wb.create_sheet("6_Month_End_Close")
    
    # Title
    ws['A1'] = "TAB 6: MONTH-END CLOSE PACKAGE"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    
    ws['A3'] = "SCENARIO"
    set_title_style(ws['A3'])
    ws['A4'] = "It's January 31, 2026. You need to prepare Treasury's month-end close package for Accounting."
    ws['A5'] = "This includes debt balances, interest expense, lease expense, and summary metrics for leadership."
    
    ws['A7'] = "YOUR TASK"
    set_title_style(ws['A7'])
    ws['A8'] = "1. Calculate beginning and ending debt balances for January"
    ws['A9'] = "2. Calculate total interest expense for the month"
    ws['A10'] = "3. Calculate total lease expense for the month"
    ws['A11'] = "4. Prepare journal entry format (Debit/Credit)"
    ws['A12'] = "5. Build an executive summary dashboard"
    ws['A13'] = "6. Verify all numbers reconcile to previous tabs"
    
    ws['A15'] = "DEBT PORTFOLIO SUMMARY"
    set_title_style(ws['A15'])
    
    # Debt summary section
    summary_labels = [
        ('Beginning Debt Balance (Jan 1):', '[From Tab 2 amortization schedule]'),
        ('Principal Payments Made:', '[Sum of principal paid in January]'),
        ('Ending Debt Balance (Jan 31):', '[Beginning Balance - Principal Payments]'),
        ('', ''),
        ('Interest Expense - January:', '[From Tab 2 - interest paid this month]'),
        ('Operating Lease Expense - January:', '[From Tab 1 - sum of lease payments]'),
    ]
    
    row = 16
    for label, hint in summary_labels:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = hint
        ws[f'B{row}'].font = Font(italic=True, color="7F7F7F")
        row += 1
    
    ws['A23'] = "JOURNAL ENTRIES (Build This)"
    set_title_style(ws['A23'])
    
    # JE headers
    je_headers = ['Account', 'Description', 'Debit', 'Credit']
    for col_num, header in enumerate(je_headers, 1):
        cell = ws.cell(row=24, column=col_num)
        cell.value = header
        set_header_style(cell)
    
    # JE template
    je_data = [
        ['Interest Expense', 'January 2026 debt interest', '[Amount from above]', ''],
        ['Cash', 'Payment of interest', '', '[Same amount]'],
        ['', '', '', ''],
        ['Lease Expense', 'January 2026 operating leases', '[Amount from above]', ''],
        ['Cash', 'Payment of lease obligations', '', '[Same amount]'],
        ['', '', '', ''],
        ['Debt - Long Term', 'Principal payment on aircraft loans', '', '[Amount from above]'],
        ['Cash', 'Payment of principal', '', '[Same amount]'],
    ]
    
    row = 25
    for data in je_data:
        if len(data) == 4:
            ws[f'A{row}'] = data[0]
            ws[f'B{row}'] = data[1]
            ws[f'C{row}'] = data[2]
            ws[f'D{row}'] = data[3]
            if data[2]:
                ws[f'C{row}'].font = Font(italic=True, color="7F7F7F")
            if data[3]:
                ws[f'D{row}'].font = Font(italic=True, color="7F7F7F")
        row += 1
    
    ws['A34'] = "EXECUTIVE SUMMARY DASHBOARD"
    set_title_style(ws['A34'])
    
    dashboard_metrics = [
        ('', ''),
        ('KEY METRICS - JANUARY 2026', ''),
        ('', ''),
        ('Total Debt Outstanding:', '[Ending debt balance]'),
        ('Debt-to-Assets Ratio:', '[If total assets = $500M, calculate ratio]'),
        ('', ''),
        ('Monthly Cash Outflow:', '[Total of all January payments]'),
        ('Annual Run Rate:', '[Monthly * 12]'),
        ('', ''),
        ('Interest Coverage:', '[Operating income / Interest expense]'),
        ('Average Interest Rate:', '[Total interest / Avg debt balance]'),
        ('', ''),
        ('Fleet Count:', '5 aircraft'),
        ('Owned (Finance Lease):', '2 aircraft'),
        ('Leased (Operating):', '3 aircraft'),
    ]
    
    row = 35
    for label, value in dashboard_metrics:
        ws[f'A{row}'] = label
        if label and not label.startswith('KEY'):
            ws[f'A{row}'].font = Font(bold=True)
        elif 'KEY METRICS' in label:
            ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
        ws[f'B{row}'] = value
        if value and value.startswith('['):
            ws[f'B{row}'].font = Font(italic=True, color="7F7F7F")
        row += 1
    
    ws['A52'] = "RECONCILIATION CHECKLIST"
    set_title_style(ws['A52'])
    
    checklist = [
        '☐ Debt balance agrees to amortization schedule (Tab 2)',
        '☐ Interest expense agrees to amortization schedule (Tab 2)',
        '☐ Lease expense agrees to fleet portfolio (Tab 1)',
        '☐ Total cash outflow agrees to cash flow forecast (Tab 3)',
        '☐ All payments reconciled to bank statement (Tab 4)',
        '☐ Variance explanations documented (Tab 5)',
        '☐ Journal entries balanced (Debits = Credits)',
        '☐ Supporting documentation attached for audit',
    ]
    
    row = 53
    for item in checklist:
        ws[f'A{row}'] = item
        ws[f'A{row}'].font = Font(size=10)
        row += 1
    
    ws['A62'] = "HINTS:"
    ws['A62'].font = Font(bold=True, color="E67E22")
    ws['A63'] = "• Pull data from previous tabs using cell references"
    ws['A64'] = "• Journal entries must balance (total debits = total credits)"
    ws['A65'] = "• Use formulas for all calculations - no hard-coded numbers"
    ws['A66'] = "• Format currency cells consistently with $#,##0 format"
    ws['A67'] = "• This package would be sent to Accounting for their close process"
    
    ws['A69'] = "WHY THIS MATTERS:"
    ws['A69'].font = Font(bold=True, color="1F4E78")
    ws['A70'] = "Month-end close package demonstrates:"
    ws['A71'] = "  • Treasury's role supporting Accounting close"
    ws['A72'] = "  • Accuracy and attention to detail"
    ws['A73'] = "  • Understanding of journal entries and accounting principles"
    ws['A74'] = "  • Ability to synthesize data from multiple sources"
    ws['A75'] = "  • Communication with cross-functional teams (Finance, Accounting, Leadership)"
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

if __name__ == "__main__":
    print("Creating Alaska Airlines Treasury Practice Simulation...")
    filename = create_treasury_simulation()
    print(f"\n[SUCCESS] Workbook created successfully: {filename}")
    print("\nNext steps:")
    print("1. Open the workbook in Excel")
    print("2. Start with the Instructions tab")
    print("3. Work through exercises 1-6 sequentially")
    print("4. Refer to Treasury_Practice_Guide.md for hints")
    print("\nGood luck! This simulation will prepare you for Alaska's Treasury role.")
